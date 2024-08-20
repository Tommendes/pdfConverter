import pdfplumber
import pandas as pd
from docx import Document
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import sys
import os

def convert_pdf_to_formats(pdf_path):
    # Definir os caminhos dos arquivos de saída
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    excel_path = f"{base_name}.xlsx"
    csv_path = f"{base_name}.csv"
    txt_path = f"{base_name}.txt"

    # Abrir o arquivo PDF
    with pdfplumber.open(pdf_path) as pdf:
        tables = []
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                tables.extend(table)

    # Garantir que há dados extraídos
    if not tables:
        print("Nenhuma tabela foi encontrada no PDF.")
        return

    # Ajustar as tabelas para que todas as linhas tenham o mesmo número de colunas
    max_cols = max(len(row) for row in tables)
    normalized_tables = [row + [''] * (max_cols - len(row)) for row in tables]

    # Converter para DataFrame
    df = pd.DataFrame(normalized_tables[1:], columns=normalized_tables[0])

    # Salvar o DataFrame como um arquivo CSV
    with open(csv_path, 'w', newline='\n') as f:
        df.to_csv(f, index=False, sep=';', quotechar='"')

    # Salvar o DataFrame como um arquivo TXT
    with open(txt_path, 'w') as txt_file:
        txt_file.write(df.to_string(index=False))

    # Salvar o DataFrame como um arquivo Excel
    df.to_excel(excel_path, index=False)

    # Formatar o arquivo Excel
    wb = load_workbook(excel_path)
    ws = wb.active

    # Definir um estilo de formatação para valores monetários
    currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')

    # Aplicar o estilo a uma coluna específica (por exemplo, a terceira coluna, índice 3)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=ws.max_row):
        for cell in row:
            cell.style = currency_style

    # Salvar o arquivo Excel com a formatação
    wb.save(excel_path)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python pdfConverter.py <arquivo_pdf>")
        sys.exit(1)

    pdf_file = sys.argv[1]
    if not os.path.isfile(pdf_file):
        print(f"Arquivo não encontrado: {pdf_file}")
        sys.exit(1)

    # Chamar a função de conversão
    convert_pdf_to_formats(pdf_file)
